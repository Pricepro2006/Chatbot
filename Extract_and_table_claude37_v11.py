# extract_and_table_v11_fixed.py
import os
import shutil
import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
import argparse

# --- CONFIGURATION ---
CURRENT_DEALS_FOLDER = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\HPI\\Deal Repository\\Current Deals'
PREVIOUS_DEALS_FOLDER = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\HPI\\Deal Repository\\Previous Deals'
MASTER_FILE_PATH = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\HPI\\Deal Repository\\Master Files\\master_deals.xlsx'
CUSTOMER_COPY_PATH = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\Co-pilot Chatbot\\master_deal_cust.xlsx'
BACKUP_FOLDER = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\HPI\\Deal Repository\\Master Files\\Backups'
LOG_FILE = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\HPI\\Deal Repository\\Master Files\\Processing_Log.txt'
HEADER_SOURCE_TRACKER = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\HPI\\Deal Repository\\Master Files\\header_source.txt'
DASHBOARD_FILE = r'C:\\Users\\mrshr\\OneDrive - TDSYNNEX\\HPI\\Deal Repository\\Master Files\\dashboard.txt'
ARCHIVE_FOLDER = os.path.join(os.path.dirname(CURRENT_DEALS_FOLDER), "Archive")

os.makedirs(BACKUP_FOLDER, exist_ok=True)
os.makedirs(ARCHIVE_FOLDER, exist_ok=True)

# --- UTILITY FUNCTIONS ---
def backup_master_file():
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(BACKUP_FOLDER, f'master_deals_backup_{timestamp}.xlsx')
    if os.path.exists(MASTER_FILE_PATH):  # Only backup if file exists
        shutil.copy2(MASTER_FILE_PATH, backup_path)
        print(f"📄 Backup created: {backup_path}")
        return True
    else:
        print(f"⚠️ No backup created - master file does not exist yet")
        return False

def parse_deal_filename(filename):
    try:
        name = filename.replace("translate_quote_", "").replace("_all.xlsx", "")
        parts = name.split("_v")
        return parts[0], int(parts[1])
    except Exception:
        return None, None

def initialize_master_file():
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create all necessary sheets
    for name in ["Deals", "Bundles", "Summary", "Previous Deals", "Master Deal History", "System_Info"]:
        wb.create_sheet(name)
    
    # Initialize Summary tab with headers
    summary_ws = wb["Summary"]
    summary_ws.append(["DealBase", "Deal Name", "Version", "Customer", "Product Numbers?", "Bundles?"])
    
    # Initialize Previous Deals tab with same headers
    prev_deals_ws = wb["Previous Deals"]
    prev_deals_ws.append(["DealBase", "Deal Name", "Version", "Customer", "Product Numbers?", "Bundles?", "Archived Date"])
    
    # Initialize Master Deal History tab
    history_ws = wb["Master Deal History"]
    history_ws.append(["DealBase", "Version", "Timestamp", "Status"])
    
    # Initialize Deals and Bundles tabs with basic headers
    deals_ws = wb["Deals"]
    deals_ws.append(["DealBase", "Version", "Customer", "Product Family", "Product Number", "Dealer Net Price", "Remaining Qty"])
    
    bundles_ws = wb["Bundles"]
    bundles_ws.append(["DealBase", "Version", "Customer", "Bundle Description", "Component 1", "Component 2"])
    
    wb.save(MASTER_FILE_PATH)
    print("✅ Initialized master_deals.xlsx with all required sheets")

def check_true_duplicates(current_files, previous_files):
    """
    Identify deals that exist with the same dealbase AND version in both folders
    Returns a list of (dealbase, version, current_filename, prev_filename) tuples
    """
    true_duplicates = []
    
    # Build dictionaries mapping (dealbase, version) to filename for both folders
    current_deals = {}
    for filename, (dealbase, version) in current_files.items():
        if dealbase and version:
            current_deals[(dealbase, version)] = filename
    
    previous_deals = {}
    for filename, (dealbase, version) in previous_files.items():
        if dealbase and version:
            previous_deals[(dealbase, version)] = filename
    
    # Find exact duplicates (same dealbase AND version)
    for key in set(current_deals.keys()) & set(previous_deals.keys()):
        dealbase, version = key
        true_duplicates.append((dealbase, version, current_deals[key], previous_deals[key]))
    
    return true_duplicates

def resolve_duplicate_deals():
    """
    Handle duplicate deal files according to these rules:
    1. If same version in both folders and it's the most current, keep in Current Deals, move from Previous to Archive
    2. If Previous Deals has version that's exactly one less than Current Deals, that's correct (keep both)
    3. All other cases (older versions, multiple versions, etc.) go to Archive
    """
    # Build dictionaries mapping dealbase to a dict of version:filename
    current_files = {}
    for filename in os.listdir(CURRENT_DEALS_FOLDER):
        if filename.lower().endswith(".xlsx") and not filename.startswith('~$'):
            deal_base, version = parse_deal_filename(filename)
            if deal_base:
                if deal_base not in current_files:
                    current_files[deal_base] = {}
                current_files[deal_base][version] = filename
    
    previous_files = {}
    for filename in os.listdir(PREVIOUS_DEALS_FOLDER):
        if filename.lower().endswith(".xlsx") and not filename.startswith('~$'):
            deal_base, version = parse_deal_filename(filename)
            if deal_base:
                if deal_base not in previous_files:
                    previous_files[deal_base] = {}
                previous_files[deal_base][version] = filename
    
    # Find overlapping dealbases
    overlapping_dealbases = set(current_files.keys()) & set(previous_files.keys())
    
    moves_needed = []  # (source, destination, reason) tuples
    
    for deal_base in overlapping_dealbases:
        # Get all versions for this dealbase
        current_versions = current_files[deal_base]
        previous_versions = previous_files[deal_base]
        
        # Find highest version in Current Deals
        current_max_version = max(current_versions.keys()) if current_versions else 0
        
        # Process each version in Previous Deals
        for prev_version, prev_filename in list(previous_versions.items()):  # Use list to avoid dictionary changed during iteration
            if prev_version in current_versions:
                # Same version in both folders
                if prev_version == current_max_version:
                    # It's the highest version - keep in Current, move from Previous to Archive
                    moves_needed.append((
                        os.path.join(PREVIOUS_DEALS_FOLDER, prev_filename),
                        os.path.join(ARCHIVE_FOLDER, prev_filename),
                        f"Same version {prev_version} in both folders (keeping in Current)"
                    ))
                else:
                    # Not the highest version - move both to Archive
                    moves_needed.append((
                        os.path.join(PREVIOUS_DEALS_FOLDER, prev_filename),
                        os.path.join(ARCHIVE_FOLDER, prev_filename),
                        f"Duplicate non-current version {prev_version}"
                    ))
                    moves_needed.append((
                        os.path.join(CURRENT_DEALS_FOLDER, current_versions[prev_version]),
                        os.path.join(ARCHIVE_FOLDER, current_versions[prev_version]),
                        f"Duplicate non-current version {prev_version}"
                    ))
            elif prev_version == current_max_version - 1:
                # Previous has exactly one version lower - this is correct
                pass
            else:
                # Previous has some other version - move to Archive
                moves_needed.append((
                    os.path.join(PREVIOUS_DEALS_FOLDER, prev_filename),
                    os.path.join(ARCHIVE_FOLDER, prev_filename),
                    f"Non-sequential version (Current has {current_max_version}, Previous has {prev_version})"
                ))
        
        # Process Current Deals versions too
        for curr_version, curr_filename in list(current_versions.items()):  # Use list to avoid dictionary changed during iteration
            if curr_version != current_max_version and curr_version != current_max_version - 1:
                # Current Deals has older versions that aren't the current or previous - move to Archive
                moves_needed.append((
                    os.path.join(CURRENT_DEALS_FOLDER, curr_filename),
                    os.path.join(ARCHIVE_FOLDER, curr_filename),
                    f"Outdated version {curr_version} in Current (max is {current_max_version})"
                ))
    
    return overlapping_dealbases, moves_needed

def clean_duplicate_deals(dry_run=False):
    """
    Execute file moves to resolve duplicated deals
    If dry_run is True, only print what would be done without moving files
    """
    overlapping_dealbases, moves_needed = resolve_duplicate_deals()
    
    print(f"Found {len(overlapping_dealbases)} dealbase(s) that appear in both folders")
    print(f"Need to perform {len(moves_needed)} file move(s) to resolve issues")
    
    if dry_run:
        print("\nDRY RUN - Not actually moving files")
        for source, destination, reason in moves_needed:
            print(f"Would move: {os.path.basename(source)} -> {os.path.dirname(destination)}")
            print(f"  Reason: {reason}")
        return 0
    
    moved_count = 0
    for source, destination, reason in moves_needed:
        try:
            print(f"Moving: {os.path.basename(source)} -> {os.path.dirname(destination)}")
            print(f"  Reason: {reason}")
            shutil.move(source, destination)
            moved_count += 1
        except Exception as e:
            print(f"❌ Error moving {os.path.basename(source)}: {e}")
    
    print(f"✅ Completed {moved_count} of {len(moves_needed)} file moves")
    return moved_count

def generate_dashboard(current_files, previous_files, master_deals_info):
    """Generate a dashboard with checks and comparisons"""
    with open(DASHBOARD_FILE, 'w', encoding='utf-8') as f:
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        f.write(f"# MASTER DEALS DASHBOARD - Generated on {now}\n\n")
        
        # Summary counts
        f.write(f"## SUMMARY COUNTS\n")
        f.write(f"Current Deals Files: {len(current_files)}\n")
        f.write(f"Previous Deals Files: {len(previous_files)}\n")
        f.write(f"Total Unique DealBases in Master File: {len(master_deals_info['all_dealbases'])}\n")
        f.write(f"DealBases in Summary Tab (Current): {len(master_deals_info['summary_dealbases'])}\n")
        f.write(f"DealBases in Previous Deals Tab: {len(master_deals_info['previous_dealbases'])}\n")
        f.write(f"DealBases in Master History Tab: {len(master_deals_info['history_dealbases'])}\n\n")
        
        # Consistency checks
        f.write(f"## CONSISTENCY CHECKS\n")
        
        # Check 1: Files in Current Deals folder but not in Summary tab
        missing_in_summary = [f for f, (db, _) in current_files.items() 
                            if db and db not in master_deals_info['summary_dealbases']]
        f.write(f"Files in Current Deals folder but not in Summary tab: {len(missing_in_summary)}\n")
        if missing_in_summary:
            f.write("    " + ", ".join(missing_in_summary[:5]))
            if len(missing_in_summary) > 5:
                f.write(f" and {len(missing_in_summary) - 5} more...")
            f.write("\n")
        
        # Check 2: Files in Previous Deals folder but not in Previous Deals tab
        missing_in_prev_tab = [f for f, (db, _) in previous_files.items() 
                              if db and db not in master_deals_info['previous_dealbases']]
        f.write(f"Files in Previous Deals folder but not in Previous Deals tab: {len(missing_in_prev_tab)}\n")
        if missing_in_prev_tab:
            f.write("    " + ", ".join(missing_in_prev_tab[:5]))
            if len(missing_in_prev_tab) > 5:
                f.write(f" and {len(missing_in_prev_tab) - 5} more...")
            f.write("\n")
        
        # Check 3: Deals in Summary but files not found in Current Deals folder
        deals_in_summary_not_in_current = [db for db in master_deals_info['summary_dealbases'] 
                                          if db not in [d for _, (d, _) in current_files.items() if d]]
        f.write(f"DealBases in Summary tab but files not found in Current Deals folder: {len(deals_in_summary_not_in_current)}\n")
        if deals_in_summary_not_in_current:
            f.write("    " + ", ".join(deals_in_summary_not_in_current[:5]))
            if len(deals_in_summary_not_in_current) > 5:
                f.write(f" and {len(deals_in_summary_not_in_current) - 5} more...")
            f.write("\n")
        
        # Check 4: Previous deals in tab but files not found in Previous Deals folder
        prev_in_tab_not_in_folder = [db for db in master_deals_info['previous_dealbases'] 
                                    if db not in [d for _, (d, _) in previous_files.items() if d]]
        f.write(f"DealBases in Previous Deals tab but files not found in Previous Deals folder: {len(prev_in_tab_not_in_folder)}\n")
        if prev_in_tab_not_in_folder:
            f.write("    " + ", ".join(prev_in_tab_not_in_folder[:5]))
            if len(prev_in_tab_not_in_folder) > 5:
                f.write(f" and {len(prev_in_tab_not_in_folder) - 5} more...")
            f.write("\n")
        
        # Check 5A: DealBases in both folders (regardless of version)
        current_deal_bases = set(db for _, (db, _) in current_files.items() if db)
        previous_deal_bases = set(db for _, (db, _) in previous_files.items() if db)
        overlapping_dealbases = current_deal_bases & previous_deal_bases
        
        f.write(f"DealBases found in both Current and Previous Deals folders: {len(overlapping_dealbases)}\n")
        if overlapping_dealbases:
            f.write("    " + ", ".join(list(overlapping_dealbases)[:5]))
            if len(overlapping_dealbases) > 5:
                f.write(f" and {len(overlapping_dealbases) - 5} more...")
            f.write("\n")
        
        # Check 5B: EXACT duplicates (same dealbase AND version)
        true_duplicates = check_true_duplicates(current_files, previous_files)
        f.write(f"Files with identical DealBase AND Version in both folders: {len(true_duplicates)}\n")
        if true_duplicates:
            examples = [f"{db} v{ver}" for db, ver, _, _ in true_duplicates[:5]]
            f.write("    " + ", ".join(examples))
            if len(true_duplicates) > 5:
                f.write(f" and {len(true_duplicates) - 5} more...")
            f.write("\n\n")
        
        # Check 6: Master History completeness
        not_in_history = []
        for db in master_deals_info['all_dealbases']:
            if db not in master_deals_info['history_dealbases']:
                not_in_history.append(db)
        
        f.write(f"DealBases not found in Master Deal History tab: {len(not_in_history)}\n")
        if not_in_history:
            f.write("    " + ", ".join(not_in_history[:5]))
            if len(not_in_history) > 5:
                f.write(f" and {len(not_in_history) - 5} more...")
            f.write("\n\n")
        
        # Overall health assessment
        same_dealbase_count = len(overlapping_dealbases)
        true_duplicate_count = len(true_duplicates)
        other_issues = len(missing_in_summary) + len(missing_in_prev_tab) + len(deals_in_summary_not_in_current) + len(prev_in_tab_not_in_folder) + len(not_in_history)
        
        total_issues = other_issues
        if true_duplicate_count > 0:
            total_issues += true_duplicate_count  # Only count exact duplicates as issues
        
        if total_issues == 0:
            f.write("## HEALTH STATUS: EXCELLENT ✅\n")
            f.write("All files are properly tracked in the master file and all tabs are consistent with folder contents.\n")
        elif total_issues < 5:
            f.write("## HEALTH STATUS: GOOD ⚠️\n")
            f.write(f"Found {total_issues} minor inconsistencies that should be addressed.\n")
        else:
            f.write("## HEALTH STATUS: NEEDS ATTENTION ❌\n")
            if true_duplicate_count > 0:
                f.write(f"Found {true_duplicate_count} files with identical versions in both folders that require attention.\n")
            if other_issues > 0:
                f.write(f"Found {other_issues} other inconsistencies that require attention.\n")

# --- MAIN FUNCTION ---
def main(test_mode=False):
    print("🚀 Starting Extract v11 with Previous Deals Support and Duplicate Resolution...")

    if not os.path.exists(MASTER_FILE_PATH):
        initialize_master_file()
    else:
        try:
            # Check if required tabs exist, create if they don't
            wb = load_workbook(MASTER_FILE_PATH)
            
            # Ensure Previous Deals tab exists
            if "Previous Deals" not in wb.sheetnames:
                prev_deals_ws = wb.create_sheet("Previous Deals")
                prev_deals_ws.append(["DealBase", "Deal Name", "Version", "Customer", "Product Numbers?", "Bundles?", "Archived Date"])
                print("✅ Added Previous Deals tab to existing master file")
            
            # Ensure Master Deal History tab has status column
            if "Master Deal History" in wb.sheetnames:
                history_ws = wb["Master Deal History"]
                if history_ws.max_column < 4:  # If Status column doesn't exist
                    if history_ws.max_row > 0:  # If there are any rows
                        # Add Status header
                        history_ws.cell(row=1, column=4).value = "Status"
                        # Add "Current" for existing entries
                        for row in range(2, history_ws.max_row + 1):
                            history_ws.cell(row=row, column=4).value = "Current"
            
            wb.save(MASTER_FILE_PATH)
            wb.close()
        except Exception as e:
            print(f"⚠️ Warning: Error checking master file structure: {e}")
            print("Creating backup and continuing...")

    backup_master_file()

    # Get files from both Current and Previous Deals folders
    current_files = {
        f: parse_deal_filename(f) for f in os.listdir(CURRENT_DEALS_FOLDER)
        if f.lower().endswith(".xlsx") and not f.startswith('~$')
    }
    
    previous_files = {
        f: parse_deal_filename(f) for f in os.listdir(PREVIOUS_DEALS_FOLDER)
        if f.lower().endswith(".xlsx") and not f.startswith('~$')
    }
    
    print(f"📄 Found {len(current_files)} Current Deal files and {len(previous_files)} Previous Deal files")

    if test_mode:
        # Limit files in test mode
        current_files_list = list(current_files.keys())[:15]
        previous_files_list = list(previous_files.keys())[:15]
        current_files = {k: current_files[k] for k in current_files_list}
        previous_files = {k: previous_files[k] for k in previous_files_list}

    locked_deals_headers = None
    locked_bundles_headers = None
    locked_deals_source = ""
    locked_bundles_source = ""

    new_deals_rows = []
    new_bundles_rows = []
    new_summary_rows = []
    new_previous_deals_rows = []
    new_history_rows = []
    
    # Load existing history entries to avoid duplicates
    existing_history_entries = set()
    try:
        history_df = pd.read_excel(MASTER_FILE_PATH, sheet_name="Master Deal History", engine="openpyxl")
        for _, row in history_df.iterrows():
            if not pd.isna(row["DealBase"]) and not pd.isna(row["Version"]):
                entry_key = f"{row['DealBase']}_{row['Version']}"
                existing_history_entries.add(entry_key)
    except Exception as e:
        print(f"Note: Could not load existing history entries: {e}")

    with open(LOG_FILE, 'w', encoding='utf-8') as log_file:
        # Process Current Deals
        print("Processing Current Deals...")
        for filename, (deal_base, version) in tqdm(current_files.items(), desc="Processing Current Deals"):
            if not deal_base:
                continue

            filepath = os.path.join(CURRENT_DEALS_FOLDER, filename)
            try:
                wb = load_workbook(filepath, data_only=True)
                sheetnames = [s.lower() for s in wb.sheetnames]

                customer = "Unknown Customer"
                if 'product numbers' in sheetnames:
                    customer_cell = wb[wb.sheetnames[sheetnames.index('product numbers')]]['B4'].value
                elif 'bundles' in sheetnames:
                    customer_cell = wb[wb.sheetnames[sheetnames.index('bundles')]]['B4'].value
                else:
                    customer_cell = None

                if customer_cell and 'for ' in str(customer_cell):
                    customer = str(customer_cell).split('for ')[-1].strip()

                # --- Product Numbers ---
                has_products = False
                if 'product numbers' in sheetnames:
                    has_products = True
                    ws = wb[wb.sheetnames[sheetnames.index('product numbers')]]
                    try:
                        if not locked_deals_headers:
                            # Get headers from row 8
                            headers = []
                            for cell in ws[8]:
                                if cell.value:
                                    headers.append(cell.value)
                            
                            if headers:  # Only set if we found headers
                                locked_deals_headers = headers + ["DealBase", "Version", "Customer"]
                                locked_deals_source = filename
                                print(f"✅ Captured Product Numbers headers from {filename}")
                            else:
                                log_file.write(f"⚠️ No headers found in Product Numbers tab for {filename}\n")
                                
                        # Process data rows starting from row 10
                        for row in ws.iter_rows(min_row=10, values_only=True):
                            if any(cell is not None and str(cell).strip() != "" for cell in row):
                                if locked_deals_headers:
                                    # Use locked_deals_headers length as a guide
                                    data_length = len(locked_deals_headers) - 3  # minus DealBase, Version, Customer
                                    dynamic_data = list(row)[:data_length]
                                    # Fill with None if row has fewer columns than headers
                                    while len(dynamic_data) < data_length:
                                        dynamic_data.append(None)
                                    new_data_row = dynamic_data + [deal_base, version, customer]
                                    new_deals_rows.append(new_data_row)
                                else:
                                    # Fallback if no headers set yet
                                    new_deals_rows.append(list(row) + [deal_base, version, customer])
                    except Exception as e:
                        log_file.write(f"❌ Error processing Product Numbers for {filename}: {str(e)}\n")

                # --- Bundles ---
                has_bundles = False
                if 'bundles' in sheetnames:
                    has_bundles = True
                    ws = wb[wb.sheetnames[sheetnames.index('bundles')]]
                    try:
                        if not locked_bundles_headers:
                            # Get headers from row 8
                            headers = []
                            for cell in ws[8]:
                                if cell.value:
                                    headers.append(cell.value)
                            
                            if headers:  # Only set if we found headers
                                locked_bundles_headers = headers + ["DealBase", "Version", "Customer"]
                                locked_bundles_source = filename
                                print(f"✅ Captured Bundles headers from {filename}")
                            else:
                                log_file.write(f"⚠️ No headers found in Bundles tab for {filename}\n")
                        
                        # Process data rows starting from row 9
                        for row in ws.iter_rows(min_row=9, values_only=True):
                            if any(cell is not None and str(cell).strip() != "" for cell in row):
                                if locked_bundles_headers:
                                    # Use locked_bundles_headers length as a guide
                                    data_length = len(locked_bundles_headers) - 3  # minus DealBase, Version, Customer
                                    dynamic_data = list(row)[:data_length]
                                    # Fill with None if row has fewer columns than headers
                                    while len(dynamic_data) < data_length:
                                        dynamic_data.append(None)
                                    new_data_row = dynamic_data + [deal_base, version, customer]
                                    new_bundles_rows.append(new_data_row)
                                else:
                                    # Fallback if no headers set yet
                                    new_bundles_rows.append(list(row) + [deal_base, version, customer])
                    except Exception as e:
                        log_file.write(f"❌ Error processing Bundles for {filename}: {str(e)}\n")

                # Add to Summary (Current Deals only)
                new_summary_rows.append([
                    deal_base,
                    f"{deal_base} v.{version}",
                    version,
                    customer,
                    'Y' if has_products else '',
                    'Y' if has_bundles else ''
                ])
                
                # Add to Master Deal History if not already there
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                entry_key = f"{deal_base}_{version}"
                if entry_key not in existing_history_entries:
                    new_history_rows.append([deal_base, version, timestamp, "Current"])
                    existing_history_entries.add(entry_key)
                
                wb.close()

            except Exception as e:
                log_file.write(f"❌ Current Deal {filename} failed: {str(e)}\n")

        # Process Previous Deals (for Previous Deals tab)
        print("Processing Previous Deals...")
        for filename, (deal_base, version) in tqdm(previous_files.items(), desc="Processing Previous Deals"):
            if not deal_base:
                continue

            filepath = os.path.join(PREVIOUS_DEALS_FOLDER, filename)
            try:
                wb = load_workbook(filepath, data_only=True)
                sheetnames = [s.lower() for s in wb.sheetnames]

                customer = "Unknown Customer"
                if 'product numbers' in sheetnames:
                    customer_cell = wb[wb.sheetnames[sheetnames.index('product numbers')]]['B4'].value
                elif 'bundles' in sheetnames:
                    customer_cell = wb[wb.sheetnames[sheetnames.index('bundles')]]['B4'].value
                else:
                    customer_cell = None

                if customer_cell and 'for ' in str(customer_cell):
                    customer = str(customer_cell).split('for ')[-1].strip()

                has_products = 'product numbers' in sheetnames
                has_bundles = 'bundles' in sheetnames

                # Add to Previous Deals tab
                archived_date = datetime.datetime.now().strftime('%Y-%m-%d')
                new_previous_deals_rows.append([
                    deal_base,
                    f"{deal_base} v.{version}",
                    version,
                    customer,
                    'Y' if has_products else '',
                    'Y' if has_bundles else '',
                    archived_date
                ])
                
                # Add to Master Deal History if not already there
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                entry_key = f"{deal_base}_{version}"
                if entry_key not in existing_history_entries:
                    new_history_rows.append([deal_base, version, timestamp, "Previous"])
                    existing_history_entries.add(entry_key)
                
                wb.close()

            except Exception as e:
                log_file.write(f"❌ Previous Deal {filename} failed: {str(e)}\n")

    # --- Write Headers + Data ---
    try:
        wb = load_workbook(MASTER_FILE_PATH)

        # Deals tab (current deals detailed data)
        deals_ws = wb["Deals"]
        deals_ws.delete_rows(1, deals_ws.max_row)
        if locked_deals_headers:
            deals_ws.append(locked_deals_headers)
            for row in new_deals_rows:
                deals_ws.append(row)
        else:
            # Fallback to basic headers if no headers were found
            deals_ws.append(["DealBase", "Version", "Customer", "Product Family", "Product Number", "Dealer Net Price", "Remaining Qty"])
            print("⚠️ Using default Product Numbers headers (no headers found in files)")

        # Bundles tab (current deals bundles)
        bundles_ws = wb["Bundles"]
        bundles_ws.delete_rows(1, bundles_ws.max_row)
        if locked_bundles_headers:
            bundles_ws.append(locked_bundles_headers)
            for row in new_bundles_rows:
                bundles_ws.append(row)
        else:
            # Fallback to basic headers if no headers were found
            bundles_ws.append(["DealBase", "Version", "Customer", "Bundle Description", "Component 1", "Component 2"])
            print("⚠️ Using default Bundles headers (no headers found in files)")

        # Summary tab (current deals summary)
        summary_ws = wb["Summary"]
        summary_ws.delete_rows(1, summary_ws.max_row)
        summary_ws.append(["DealBase", "Deal Name", "Version", "Customer", "Product Numbers?", "Bundles?"])
        for row in new_summary_rows:
            summary_ws.append(row)

        # Previous Deals tab (previous deals summary)
        prev_deals_ws = wb["Previous Deals"]
        prev_deals_ws.delete_rows(1, prev_deals_ws.max_row)
        prev_deals_ws.append(["DealBase", "Deal Name", "Version", "Customer", "Product Numbers?", "Bundles?", "Archived Date"])
        for row in new_previous_deals_rows:
            prev_deals_ws.append(row)

        # History tab (append new entries)
        history_ws = wb["Master Deal History"]
        if history_ws.max_row <= 1:  # If only header or empty
            history_ws.delete_rows(1, history_ws.max_row)
            history_ws.append(["DealBase", "Version", "Timestamp", "Status"])
        
        # Add new history entries
        for row in new_history_rows:
            history_ws.append(row)

        wb.save(MASTER_FILE_PATH)
        wb.close()
        print("✅ Successfully saved all data to master file")
    except Exception as e:
        print(f"❌ Error writing to master file: {e}")
        return

    with open(HEADER_SOURCE_TRACKER, 'w') as f:
        f.write(f"Deals Header Source: {locked_deals_source}\n")
        f.write(f"Bundles Header Source: {locked_bundles_source}\n")

    # Create a customer copy if requested
    try:
        if os.path.exists(MASTER_FILE_PATH):
            shutil.copy2(MASTER_FILE_PATH, CUSTOMER_COPY_PATH)
            print(f"✅ Created customer copy at {CUSTOMER_COPY_PATH}")
    except Exception as e:
        print(f"Note: Could not create customer copy: {e}")

    # Collect information for dashboard
    master_deals_info = {
        'summary_dealbases': set(),
        'previous_dealbases': set(),
        'history_dealbases': set(),
        'all_dealbases': set()
    }
    
    try:
        wb = load_workbook(MASTER_FILE_PATH, read_only=True)
        
        # Get current dealbases from Summary tab
        ws = wb["Summary"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # DealBase
                master_deals_info['summary_dealbases'].add(str(row[0]))
                master_deals_info['all_dealbases'].add(str(row[0]))
        
        # Get previous dealbases from Previous Deals tab
        ws = wb["Previous Deals"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # DealBase column
                master_deals_info['previous_dealbases'].add(str(row[0]))
                master_deals_info['all_dealbases'].add(str(row[0]))
        
        # Get history dealbases
        ws = wb["Master Deal History"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # DealBase column
                master_deals_info['history_dealbases'].add(str(row[0]))
        
        wb.close()
        
        # Generate dashboard
        generate_dashboard(current_files, previous_files, master_deals_info)
        print("📊 Dashboard generated with consistency checks")
    except Exception as e:
        print(f"⚠️ Warning: Error generating dashboard: {e}")

    print("✅ Master file updated with both Current and Previous Deals")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process deal files and update master file")
    parser.add_argument("--clean-duplicates", action="store_true", 
                       help="Clean duplicates between Current and Previous Deals folders")
    parser.add_argument("--dry-run", action="store_true", 
                       help="Show what cleaning would do without making changes")
    parser.add_argument("--test", action="store_true", 
                       help="Run in test mode with limited files")
    args = parser.parse_args()
    
    if args.clean_duplicates:
        clean_duplicate_deals(dry_run=args.dry_run)
        if args.dry_run:
            print("Exiting after dry run")
            exit(0)
    
    main(test_mode=args.test)