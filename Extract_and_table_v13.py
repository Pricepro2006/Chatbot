# Extract_and_table_v13.py — parquet‑free, gzip‑CSV cache + same two‑version logic
"""
Changelog vs v12
----------------
1. 🆕 **Parquet -> CSV.GZ cache**
   * Removes dependency on *pyarrow* / *fastparquet*.
   * Uses `pandas.read_csv` / `to_csv(..., compression="gzip")`.
2. Graceful read‑back: if a cached CSV is corrupt or missing, falls back to re‑loading
   the original workbook and overwriting the cache.
3. Minor: factored out `read_cache()` / `write_cache()` helpers.

Everything else (latest + previous version retention, summary trimming, backup, history)
remains unchanged.
"""

import os, glob, shutil, datetime, argparse, pandas as pd
from openpyxl import load_workbook, Workbook
from tqdm import tqdm

BASE   = os.path.join(os.path.expanduser('~'), 'OneDrive - TDSYNNEX', 'HPI', 'Deal Repository')
CUR    = os.path.join(BASE, 'Current Deals')
PREV   = os.path.join(BASE, 'Previous Deals')
MSF    = os.path.join(BASE, 'Master Files', 'master_deals.xlsx')
BACK   = os.path.join(BASE, 'Master Files', 'Backups')
CACHE  = os.path.join(BASE, 'Master Files', 'CsvCache')  # renamed for clarity
HDRSRC = os.path.join(BASE, 'Master Files', 'header_source.txt')
LOG    = os.path.join(BASE, 'Master Files', 'Processing_Log.txt')

os.makedirs(BACK,  exist_ok=True)
os.makedirs(CACHE, exist_ok=True)

# ───────────────────────────────── helpers ────────────────────────────────────

def backup_master():
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    if os.path.exists(MSF):
        shutil.copy2(MSF, os.path.join(BACK, f'master_deals_backup_{ts}.xlsx'))


def parse_name(fname):
    core = fname.replace('translate_quote_', '').replace('_all.xlsx', '')
    try:
        base, ver = core.split('_v')
        return base, int(ver)
    except ValueError:
        return None, None


def keep_two_newest(df):
    idx = (
        df.sort_values(['DealBase', 'Version'])
          .groupby('DealBase')
          .tail(2)
          .index
    )
    return df.loc[idx].reset_index(drop=True)

# --- cache helpers -----------------------------------------------------------

def cache_path(fname, kind):  # kind in {'deals','bundles'}
    return os.path.join(CACHE, f'{fname}_{kind}.csv.gz')


def write_cache(df, path):
    if df.empty:
        return
    df.to_csv(path, index=False, compression='gzip')


def read_cache(path):
    try:
        return pd.read_csv(path, compression='gzip') if os.path.exists(path) else None
    except Exception:
        # corrupted cache → ignore
        return None

# ─────────────────────────────────── main ─────────────────────────────────────

def main(test=False):
    # Ensure master workbook exists ------------------------------------------------
    if not os.path.exists(MSF):
        wb = Workbook(); wb.remove(wb['Sheet'])
        for s in ['Deals', 'Bundles', 'Summary', 'Master Deal History', 'System_Info']:
            wb.create_sheet(s)
        wb.save(MSF)

    backup_master()

    # Gather files -----------------------------------------------------------------
    FILES = (
        glob.glob(os.path.join(CUR,  'translate_quote_*_v*_*.xlsx')) +
        glob.glob(os.path.join(PREV, 'translate_quote_*_v*_*.xlsx'))
    )
    FILES = sorted(set(os.path.basename(f) for f in FILES))
    if test:
        FILES = FILES[:30]

    deals_rows, bundles_rows, summary_rows, hist_rows = [], [], [], []
    lock_deals_hdr = lock_bundles_hdr = None
    lock_deals_src = lock_bundles_src = ''

    with open(LOG, 'w', encoding='utf-8') as lg:
        for fname in tqdm(FILES, desc='v13 ingest'):
            base, ver = parse_name(fname)
            if not base:
                continue

            # choose file location (Current preferred over Previous)
            x_path = os.path.join(CUR, fname) if os.path.exists(os.path.join(CUR, fname)) else os.path.join(PREV, fname)

            # try cache first ---------------------------------------------------
            df_deals   = read_cache(cache_path(fname, 'deals'))
            df_bundles = read_cache(cache_path(fname, 'bundles'))
            cust = None

            try:
                if df_deals is None and df_bundles is None:  # cache miss → parse Excel
                    wb = load_workbook(x_path, data_only=True, read_only=True)
                    sn = [s.lower() for s in wb.sheetnames]

                    # customer (basic heuristic)
                    cust = 'Unknown Customer'
                    if 'product numbers' in sn:
                        cell = wb[wb.sheetnames[sn.index('product numbers')]]['B4'].value
                    elif 'bundles' in sn:
                        cell = wb[wb.sheetnames[sn.index('bundles')]]['B4'].value
                    if cell and 'for ' in str(cell):
                        cust = str(cell).split('for ')[-1].strip()

                    # Product Numbers sheet -----------------------------------
                    if 'product numbers' in sn:
                        ws = wb[wb.sheetnames[sn.index('product numbers')]]
                        if lock_deals_hdr is None:
                            lock_deals_hdr = [c.value for c in next(ws.iter_rows(min_row=8, max_row=8)) if c.value] + ['DealBase','Version','Customer']
                            lock_deals_src = fname
                        rows = [list(r)[:len(lock_deals_hdr)-3] + [base, ver, cust] for r in ws.iter_rows(min_row=10, values_only=True) if any(r)]
                        df_deals = pd.DataFrame(rows, columns=lock_deals_hdr)
                        write_cache(df_deals, cache_path(fname, 'deals'))

                    # Bundles sheet -------------------------------------------
                    if 'bundles' in sn:
                        ws = wb[wb.sheetnames[sn.index('bundles')]]
                        if lock_bundles_hdr is None:
                            lock_bundles_hdr = [c.value for c in next(ws.iter_rows(min_row=8, max_row=8)) if c.value] + ['DealBase','Version','Customer']
                            lock_bundles_src = fname
                        rows = [list(r)[:len(lock_bundles_hdr)-3] + [base, ver, cust] for r in ws.iter_rows(min_row=9, values_only=True) if any(r)]
                        df_bundles = pd.DataFrame(rows, columns=lock_bundles_hdr)
                        write_cache(df_bundles, cache_path(fname, 'bundles'))
                    wb.close()
                else:
                    # cache hit → ensure customer value present
                    if df_deals is not None and not df_deals.empty:
                        cust = df_deals['Customer'].iloc[0]
                    elif df_bundles is not None and not df_bundles.empty:
                        cust = df_bundles['Customer'].iloc[0]
                    else:
                        cust = 'Unknown Customer'

                # accumulate rows ---------------------------------------------
                if df_deals is not None:
                    deals_rows.extend(df_deals.values.tolist())
                if df_bundles is not None:
                    bundles_rows.extend(df_bundles.values.tolist())
                summary_rows.append([base, ver, cust])
                hist_rows.append([base, ver, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

            except Exception as e:
                lg.write(f"{fname}: {e}\n")

    # DataFrames + trimming ----------------------------------------------------
    deals_df   = pd.DataFrame(deals_rows,   columns=lock_deals_hdr)
    bundles_df = pd.DataFrame(bundles_rows, columns=lock_bundles_hdr)
    deals_df   = keep_two_newest(deals_df)
    bundles_df = keep_two_newest(bundles_df)

    # Write master workbook ----------------------------------------------------
    wb = load_workbook(MSF)
    for sheet, df in {"Deals": deals_df, "Bundles": bundles_df}.items():
        ws = wb[sheet]
        ws.delete_rows(1, ws.max_row)
        for row in df.itertuples(index=False):
            ws.append(list(row))

    # Summary — active latest deals only --------------------------------------
    today = datetime.date.today()
    latest_idx = deals_df.groupby('DealBase')['Version'].idxmax()
    active_df = deals_df.loc[latest_idx]
    if 'End Date' in active_df.columns:
        active_df = active_df[pd.to_datetime(active_df['End Date'], errors='coerce').dt.date >= today]
    s_ws = wb['Summary']
    s_ws.delete_rows(1, s_ws.max_row)
    s_ws.append(['DealBase','Deal Name','Version','Customer','Product Numbers?','Bundles?'])
    for _, r in active_df.iterrows():
        s_ws.append([r['DealBase'], f"{r['DealBase']} v.{r['Version']}", r['Version'], r['Customer'], 'Y', 'Y'])

    # History (append) ---------------------------------------------------------
    h_ws = wb['Master Deal History']
    if h_ws.max_row == 1:
        h_ws.append(['DealBase','Version','Timestamp'])
    for b, v, ts in hist_rows:
        h_ws.append([b, v, ts])

    wb.save(MSF); wb.close()

    # header source tracker ----------------------------------------------------
    with open(HDRSRC, 'w') as f:
        f.write(f'Deals Header Source: {lock_deals_src}\nBundles Header Source: {lock_bundles_src}\n')

    print('✅ Extract v13 complete – parquet‑free & cached via gzip‑CSV (latest + prev version kept).')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(); parser.add_argument('--test', action='store_true'); opts = parser.parse_args()
    main(test=opts.test)
