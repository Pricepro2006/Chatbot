# brain_loader_v5.0.py - Comprehensive Synonym Brain management with hybrid loading approach
# Version 5.0 - Complete rewrite with robust error handling and improved flexibility
# Last updated: 2025-04-24

import os
import re
import csv
import json
import time
import logging
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional, Any, Union

# Advanced import handling
try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# Import the Golden Brain - always available as baseline
try:
    from golden_brain import GOLDEN_BRAIN
    HAS_GOLDEN_BRAIN = True
except ImportError:
    # Create minimal placeholder if golden_brain.py is missing
    GOLDEN_BRAIN = {}
    HAS_GOLDEN_BRAIN = False

# Import hardcoded brain as fallback
try:
    from hardcoded_brain import synonym_brain as HARDCODED_BRAIN
    HAS_HARDCODED_BRAIN = True
except ImportError:
    HARDCODED_BRAIN = {}
    HAS_HARDCODED_BRAIN = False

# Version and file path constants
VERSION = "5.0.0"
LAST_UPDATED = "2025-04-24"

# Default file paths
VARIABLES_FILE = "variable_names.xlsx"
LEARNED_SYNONYMS_FILE = "learned_synonyms.csv"
CUSTOM_SYNONYMS_FILE = "custom_synonyms.csv"
DEBUG_LOG_FILE = "brain_loader_debug.log"
BACKUP_BRAIN_JSON = "brain_backup.json"
METRICS_FILE = "brain_metrics.json"

# Configure logging with both file and console output
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("brain_loader.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("BrainLoaderV5")

# Configure performance logger
perf_logger = logging.getLogger("BrainPerformance")
perf_handler = logging.FileHandler("brain_performance.log")
perf_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
perf_logger.addHandler(perf_handler)
perf_logger.setLevel(logging.INFO)

class BrainMetrics:
    """Track and record brain performance metrics"""
    
    def __init__(self, brain_loader: Any):
        self.brain_loader = brain_loader
        self.start_time = datetime.now()
        self.load_times: Dict[str, float] = {}
        self.source_counts: Dict[str, int] = {}
        self.field_coverage: Dict[str, Dict[str, int]] = {}
        self.errors: List[Dict[str, Any]] = []
    
    def record_load_time(self, source: str, elapsed_time: float, count: int):
        """Record loading time for a source"""
        self.load_times[source] = elapsed_time
        self.source_counts[source] = count
        perf_logger.info(f"Loaded {count} items from {source} in {elapsed_time:.2f} seconds")
    
    def record_error(self, source: str, error: Exception, context: Optional[Dict[str, Any]] = None):
        """Record an error during brain loading"""
        error_data = {
            "timestamp": datetime.now().isoformat(),
            "source": source,
            "error_type": type(error).__name__,
            "error_message": str(error),
            "context": context or {}
        }
        self.errors.append(error_data)
        perf_logger.error(f"Error in {source}: {type(error).__name__}: {error}")
    
    def calculate_field_coverage(self):
        """Calculate synonym coverage per field"""
        field_counts: Dict[str, Dict[str, int]] = {}
        
        # Count by field and source
        for source, synonyms in self.brain_loader.brain_sources.items():
            for _, field in synonyms.items():
                if field not in field_counts:
                    field_counts[field] = {}
                
                if source not in field_counts[field]:
                    field_counts[field][source] = 0
                    
                field_counts[field][source] += 1
        
        self.field_coverage = field_counts
    
    def get_summary(self) -> Dict[str, Any]:
        """Get summary of brain metrics"""
        total_runtime = (datetime.now() - self.start_time).total_seconds()
        
        self.calculate_field_coverage()
        
        # Calculate total synoyms by field
        field_totals = {}
        for field, sources in self.field_coverage.items():
            field_totals[field] = sum(sources.values())
        
        return {
            "version": VERSION,
            "timestamp": datetime.now().isoformat(),
            "total_runtime": total_runtime,
            "total_synonyms": len(self.brain_loader.synonym_brain),
            "sources": self.source_counts,
            "load_times": self.load_times,
            "field_totals": field_totals,
            "field_coverage": self.field_coverage,
            "error_count": len(self.errors),
            "brain_sources": list(self.brain_loader.brain_sources.keys())
        }
    
    def save_metrics(self, file_path: str = METRICS_FILE):
        """Save metrics to JSON file"""
        try:
            with open(file_path, 'w') as f:
                json.dump(self.get_summary(), f, indent=2)
            logger.info(f"✅ Brain metrics saved to {file_path}")
            return True
        except Exception as e:
            logger.error(f"❌ Failed to save brain metrics: {e}")
            return False

class BrainLoader:
    """
    Enhanced Synonym Brain loader with a hybrid approach:
    1. Start with Golden Brain (highest priority)
    2. Enhance with Excel-based brain if available
    3. Add learned synonyms from CSV
    4. Include any custom synonyms
    
    Comprehensive error handling and fallback mechanisms ensure the system
    always has a functional brain even in case of file access failures.
    """
    
    def __init__(self):
        """Initialize the brain loader"""
        # Main brain dictionary
        self.synonym_brain: Dict[str, str] = {}
        
        # Track brain sources separately for diagnostics and metrics
        self.brain_sources: Dict[str, Dict[str, str]] = {}
        
        # Product family mapping
        self.product_family_map: Dict[str, str] = {}
        
        # Track loaded sources for reporting
        self.loaded_sources: List[str] = []
        
        # Metrics and diagnostics
        self.metrics = BrainMetrics(self)
        
        # State tracking
        self.error_count = 0
        self.last_load_time = None
        self.version = VERSION
        
        # Standard canonical fields
        self.canonical_fields = {
            "Remaining qty",
            "Dealer net price \n[USD]",
            "Product family",
            "Customer",
            "End date"
        }
    
    def normalize_sheet_name(self, name: Optional[str]) -> str:
        """
        Normalize sheet name for flexible matching by removing whitespace,
        converting to lowercase, and standardizing separators
        """
        if name is None:
            return ""
        # Handle different formats of separators and whitespace
        normalized = name.strip().lower()
        normalized = re.sub(r'[-_\s]+', ' ', normalized)
        return normalized
    
    def find_matching_sheet(self, workbook: Any, target_variations: List[str]) -> Optional[str]:
        """
        Find a sheet in the workbook that matches any of the target variations
        using both exact and fuzzy matching with detailed logging
        """
        available_sheets = workbook.sheetnames
        normalized_sheets = {self.normalize_sheet_name(name): name for name in available_sheets}
        
        # Log all available sheets for debugging
        sheet_list = ", ".join([f"'{s}'" for s in available_sheets])
        logger.info(f"Available sheets in workbook: {sheet_list}")
        
        # Step 1: Try exact match on normalized names
        for variation in target_variations:
            norm_var = self.normalize_sheet_name(variation)
            if norm_var in normalized_sheets:
                actual_name = normalized_sheets[norm_var]
                logger.info(f"✅ Found exact match: '{variation}' -> '{actual_name}'")
                return actual_name
                
        # Step 2: Try fuzzy matching (substring)
        for variation in target_variations:
            norm_var = self.normalize_sheet_name(variation)
            for norm_name, actual_name in normalized_sheets.items():
                if norm_var in norm_name or norm_name in norm_var:
                    logger.info(f"✅ Found fuzzy match: '{variation}' ~ '{norm_name}' -> '{actual_name}'")
                    return actual_name
                    
        # Step 3: Try similarity matching (shared tokens)
        for variation in target_variations:
            variation_tokens = set(self.normalize_sheet_name(variation).split())
            if not variation_tokens:
                continue
                
            best_match = None
            best_score = 0
            
            for norm_name, actual_name in normalized_sheets.items():
                sheet_tokens = set(norm_name.split())
                if not sheet_tokens:
                    continue
                    
                # Calculate intersection score
                intersection = variation_tokens.intersection(sheet_tokens)
                if intersection:
                    score = len(intersection) / max(len(variation_tokens), len(sheet_tokens))
                    if score > 0.5 and score > best_score:  # At least 50% similarity
                        best_score = score
                        best_match = actual_name
                        
            if best_match:
                logger.info(f"✅ Found token similarity match: '{variation}' ~ '{best_match}' (score: {best_score:.2f})")
                return best_match
        
        # No match found
        variations_list = ", ".join([f"'{v}'" for v in target_variations])
        logger.error(f"❌ No matching sheet found. Tried variations: {variations_list}")
        logger.error(f"❌ Available sheets: {sheet_list}")
        return None
    
    def save_brain_backup(self) -> bool:
        """Save a backup of the loaded brain for emergency recovery"""
        try:
            # Ensure the folder exists
            Path(os.path.dirname(BACKUP_BRAIN_JSON)).mkdir(parents=True, exist_ok=True)
            
            with open(BACKUP_BRAIN_JSON, 'w') as f:
                json.dump({
                    'version': VERSION,
                    'timestamp': datetime.now().isoformat(),
                    'sources': self.loaded_sources,
                    'brain': self.synonym_brain,
                    'brain_sources': {k: list(v.items()) for k, v in self.brain_sources.items()}
                }, f, indent=2)
            logger.info(f"✅ Brain backup saved to {BACKUP_BRAIN_JSON}")
            return True
        except Exception as e:
            logger.error(f"❌ Failed to save brain backup: {e}")
            return False
    
    def load_brain_backup(self) -> bool:
        """Load brain from backup if available"""
        if not os.path.exists(BACKUP_BRAIN_JSON):
            logger.warning(f"⚠️ No backup brain found at {BACKUP_BRAIN_JSON}")
            return False
            
        try:
            start_time = time.time()
            
            with open(BACKUP_BRAIN_JSON, 'r') as f:
                data = json.load(f)
                
                # Convert list items back to dict for brain_sources
                brain_sources = {}
                for source, items_list in data.get('brain_sources', {}).items():
                    brain_sources[source] = dict(items_list)
                
                self.synonym_brain = data.get('brain', {})
                self.brain_sources = brain_sources
                backup_sources = data.get('sources', [])
                backup_timestamp = data.get('timestamp', 'unknown')
                
                if self.synonym_brain:
                    self.loaded_sources.extend([f"Backup:{s}" for s in backup_sources])
                    logger.info(f"✅ Loaded {len(self.synonym_brain)} synonyms from backup (from {backup_timestamp})")
                    
                    # Record metrics
                    elapsed = time.time() - start_time
                    self.metrics.record_load_time("Backup", elapsed, len(self.synonym_brain))
                    
                    return True
                else:
                    logger.warning(f"⚠️ Backup brain was empty")
                    return False
        except Exception as e:
            logger.error(f"❌ Failed to load brain backup: {e}")
            self.metrics.record_error("Backup", e)
            return False
    
    def load_golden_brain(self) -> int:
        """
        Load the Golden Brain (highest priority synonyms)
        
        Returns:
            int: Number of synonyms loaded
        """
        if not HAS_GOLDEN_BRAIN or not GOLDEN_BRAIN:
            logger.warning("⚠️ Golden Brain not available or empty")
            return 0
            
        try:
            start_time = time.time()
            
            # Store Golden Brain separately in sources dictionary
            self.brain_sources["GoldenBrain"] = GOLDEN_BRAIN.copy()
            
            # Add to main brain
            self.synonym_brain.update(GOLDEN_BRAIN)
            
            elapsed = time.time() - start_time
            count = len(GOLDEN_BRAIN)
            
            self.loaded_sources.append(f"GoldenBrain:{count}")
            logger.info(f"✅ Loaded {count} synonyms from Golden Brain in {elapsed:.2f}s")
            
            # Record metrics
            self.metrics.record_load_time("GoldenBrain", elapsed, count)
            
            return count
        except Exception as e:
            logger.error(f"❌ Error loading Golden Brain: {e}")
            self.metrics.record_error("GoldenBrain", e)
            return 0
    
    def load_hardcoded_brain(self) -> int:
        """
        Load hardcoded brain from hardcoded_brain.py
        
        Returns:
            int: Number of synonyms loaded
        """
        if not HAS_HARDCODED_BRAIN or not HARDCODED_BRAIN:
            logger.warning("⚠️ Hardcoded Brain not available or empty")
            return 0
            
        try:
            start_time = time.time()
            
            old_size = len(self.synonym_brain)
            
            # Only store synonyms not already in the brain
            hardcoded_additions = {}
            for key, value in HARDCODED_BRAIN.items():
                if key not in self.synonym_brain:
                    hardcoded_additions[key] = value
            
            # Store additions in sources dictionary
            if hardcoded_additions:
                self.brain_sources["HardcodedBrain"] = hardcoded_additions
                
                # Add to main brain
                self.synonym_brain.update(hardcoded_additions)
            
            elapsed = time.time() - start_time
            new_size = len(self.synonym_brain)
            added = new_size - old_size
            
            if added > 0:
                self.loaded_sources.append(f"HardcodedBrain:{added}")
                logger.info(f"✅ Added {added} synonyms from Hardcoded Brain in {elapsed:.2f}s")
                
                # Record metrics
                self.metrics.record_load_time("HardcodedBrain", elapsed, added)
            else:
                logger.info("✓ No new synonyms added from Hardcoded Brain (all already present)")
            
            return added
        except Exception as e:
            logger.error(f"❌ Failed to load hardcoded brain: {e}")
            self.metrics.record_error("HardcodedBrain", e)
            return 0
    
    def load_from_excel(self, excel_path: str = VARIABLES_FILE) -> int:
        """
        Load synonyms from Excel with enhanced error handling
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            int: Number of synonyms loaded
        """
        if not HAS_OPENPYXL:
            logger.error("❌ openpyxl not available - cannot load from Excel")
            return 0
            
        if not os.path.exists(excel_path):
            logger.error(f"❌ Excel file not found: {excel_path}")
            return 0
            
        try:
            start_time = time.time()
            
            # Try to load workbook with retries
            max_retries = 3
            wb = None
            for retry in range(max_retries):
                try:
                    wb = load_workbook(excel_path, read_only=True)
                    break
                except Exception as e:
                    if retry < max_retries - 1:
                        logger.warning(f"⚠️ Retry {retry+1}/{max_retries} loading Excel file: {e}")
                        time.sleep(1)  # Wait before retry
                        continue
                    else:
                        raise
            
            if wb is None:
                logger.error(f"❌ Failed to load Excel file after {max_retries} attempts")
                return 0
            
            # Define all possible variations of the synonym brain sheet name - significantly expanded
            synonym_brain_variations = [
                # Standard formats
                "synonym brain", "synonymbrain", "synonym_brain", 
                "synonyms", "syn brain", "synbrain", "brain",
                "synonym", "brain synonyms", "brain_synonyms",
                # Project specific variants
                "nova synonyms", "field synonyms", "nova brain",
                "synonym mapping", "synonymsmap", "brainsynonyms",
                "brain map", "field map", "intent map",
                "intents", "intentions", "field brain",
                # Even more variants to catch unusual formats
                "syn", "field intent", "intent lookup",
                "field names", "field map", "field mapping",
                "mappings", "lookup", "nlp brain",
                "language brain", "term mapping", "query mapping",
                "query to field", "phrase mapping", "phrase to field",
                "intent brain", "search brain", "query brain",
                # Last resort variations
                "brain sheet", "main", "mappings"
            ]
            
            # Find the synonym brain sheet with flexible matching
            sheet_name = self.find_matching_sheet(wb, synonym_brain_variations)
            
            synonym_count = 0
            excel_synonyms = {}
            if sheet_name:
                ws = wb[sheet_name]
                
                # Count total rows for progress tracking
                try:
                    total_rows = ws.max_row
                    logger.info(f"Processing up to {total_rows} rows from sheet '{sheet_name}'")
                except:
                    # Fall back to list conversion if max_row is not reliable
                    rows = list(ws.rows)
                    total_rows = len(rows)
                    logger.info(f"Processing {total_rows} rows from sheet '{sheet_name}' (list mode)")
                
                # Process each row (skip header)
                row_num = 0
                for row in ws.iter_rows(min_row=2):
                    row_num += 1
                    
                    # Skip completely empty rows
                    if not any(cell.value for cell in row):
                        continue
                        
                    # Get target field from first column
                    target_field = row[0].value
                    if not target_field:
                        logger.warning(f"⚠️ Row {row_num+1} has no target field, skipping")
                        continue
                        
                    # Safety check - ensure target field is a string
                    target_field = str(target_field).strip()
                    
                    # Skip rows with non-canonical fields
                    if target_field not in self.canonical_fields:
                        logger.warning(f"⚠️ Row {row_num+1} has non-canonical field '{target_field}', skipping")
                        continue
                        
                    # Process synonym columns (all columns after the first)
                    for cell in row[1:]:
                        if cell.value:
                            # Add this synonym to Excel-specific dictionary
                            syn = str(cell.value).strip()
                            syn_lower = syn.lower()
                            
                            if syn_lower:  # Avoid empty strings
                                excel_synonyms[syn_lower] = target_field
                                synonym_count += 1
                    
                    # Log progress for large files
                    if row_num % 100 == 0:
                        logger.info(f"Processed {row_num}/{total_rows} rows")
                
                # Add Excel synonyms to sources dictionary
                self.brain_sources["Excel"] = excel_synonyms
                
                # Add to main brain (won't overwrite anything already there)
                for key, value in excel_synonyms.items():
                    if key not in self.synonym_brain:
                        self.synonym_brain[key] = value
                
                # Calculate actual additions
                actual_additions = sum(1 for k in excel_synonyms if k in self.synonym_brain and self.synonym_brain[k] == excel_synonyms[k])
                
                elapsed = time.time() - start_time
                
                self.loaded_sources.append(f"Excel:{actual_additions}")
                logger.info(f"✅ Loaded {synonym_count} synonyms ({actual_additions} added) from '{sheet_name}' in {elapsed:.2f}s")
                
                # Record metrics
                self.metrics.record_load_time("Excel", elapsed, actual_additions)
                
                return actual_additions
            else:
                logger.error(f"❌ Synonym Brain sheet not found in {excel_path}")
                
                # Try opening with pandas to get sheet names
                if HAS_PANDAS:
                    try:
                        xls = pd.ExcelFile(excel_path)
                        logger.info(f"Available sheets (pandas): {xls.sheet_names}")
                    except Exception as e:
                        logger.error(f"Could not read with pandas either: {e}")
                
                return 0
                
        except Exception as e:
            logger.error(f"❌ Error loading from Excel: {type(e).__name__}: {e}")
            logger.error(f"Stack trace:\n{traceback.format_exc()}")
            self.metrics.record_error("Excel", e, {"path": excel_path})
            return 0
    
    def load_from_csv(self, csv_path: str = LEARNED_SYNONYMS_FILE, source_name: str = "CSV") -> int:
        """
        Load synonyms from CSV file
        
        Args:
            csv_path: Path to CSV file
            source_name: Source name for tracking
            
        Returns:
            int: Number of synonyms loaded
        """
        if not os.path.exists(csv_path):
            # Create the file if it doesn't exist
            try:
                Path(os.path.dirname(csv_path)).mkdir(parents=True, exist_ok=True)
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Synonym", "Field", "Source", "Timestamp"])
                logger.info(f"✅ Created new {csv_path}")
                return 0  # No synonyms loaded, but file created successfully
            except Exception as e:
                logger.error(f"❌ Failed to create CSV file: {e}")
                self.metrics.record_error(source_name, e, {"path": csv_path})
                return 0
        
        # Load from existing file
        learned_count = 0
        csv_synonyms = {}
        try:
            start_time = time.time()
            
            with open(csv_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                # Skip header row but verify it exists
                header = next(reader, None)
                if not header or len(header) < 2:
                    logger.warning(f"⚠️ CSV file has invalid header: {header}")
                    
                # Process data rows
                for row in reader:
                    if len(row) >= 2 and row[0] and row[1]:
                        synonym_text, field = row[0].strip(), row[1].strip()
                        
                        # Skip rows with non-canonical fields
                        if field not in self.canonical_fields:
                            logger.debug(f"Skipping non-canonical field in CSV: '{field}'")
                            continue
                            
                        synonym_text_lower = synonym_text.lower()
                        
                        if synonym_text_lower and field:  # Avoid empty strings
                            csv_synonyms[synonym_text_lower] = field
                            learned_count += 1
            
            # Add to sources dictionary
            if source_name not in self.brain_sources:
                self.brain_sources[source_name] = {}
                
            self.brain_sources[source_name].update(csv_synonyms)
            
            # Add to main brain (won't overwrite anything already there)
            for key, value in csv_synonyms.items():
                if key not in self.synonym_brain:
                    self.synonym_brain[key] = value
                    
            # Calculate actual additions
            actual_additions = sum(1 for k in csv_synonyms if k in self.synonym_brain and self.synonym_brain[k] == csv_synonyms[k])
            
            elapsed = time.time() - start_time
            
            if actual_additions > 0:
                self.loaded_sources.append(f"{source_name}:{actual_additions}")
                logger.info(f"✅ Loaded {learned_count} synonyms ({actual_additions} added) from {csv_path} in {elapsed:.2f}s")
                
                # Record metrics
                self.metrics.record_load_time(source_name, elapsed, actual_additions)
            else:
                logger.info(f"✓ No new synonyms added from {csv_path} (all already present)")
            
            return actual_additions
        except Exception as e:
            logger.error(f"❌ Error loading from CSV {csv_path}: {type(e).__name__}: {e}")
            logger.error(f"Stack trace:\n{traceback.format_exc()}")
            self.metrics.record_error(source_name, e, {"path": csv_path})
            return 0
    
    def seed_learned_synonyms(self, csv_path: str = LEARNED_SYNONYMS_FILE) -> bool:
        """
        Add essential synonyms to learned_synonyms.csv as a backup mechanism
        with enhanced categories and better phrase coverage
        
        Args:
            csv_path: Path to CSV file
            
        Returns:
            bool: Success or failure
        """
        # Create the file if it doesn't exist
        if not os.path.exists(csv_path):
            try:
                Path(os.path.dirname(csv_path)).mkdir(parents=True, exist_ok=True)
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Synonym", "Field", "Source", "Timestamp"])
                logger.info(f"✅ Created new {csv_path}")
            except Exception as e:
                logger.error(f"❌ Failed to create CSV file: {e}")
                self.metrics.record_error("Seed", e, {"path": csv_path})
                return False
                
        # Read existing synonyms to avoid duplicates
        existing_synonyms = set()
        try:
            with open(csv_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip header
                for row in reader:
                    if len(row) >= 1:
                        existing_synonyms.add(row[0].lower())
        except Exception as e:
            logger.error(f"❌ Error reading existing synonyms: {e}")
            
        # Define essential synonyms for all categories
        # Significantly expanded from previous versions for better coverage
        essential_synonyms = [
            # Remaining qty synonyms (most problematic category)
            ["how many do we have", "Remaining qty", "Seed", datetime.now()],
            ["quantity on hand", "Remaining qty", "Seed", datetime.now()],
            ["available inventory", "Remaining qty", "Seed", datetime.now()],
            ["stock level", "Remaining qty", "Seed", datetime.now()],
            ["count of", "Remaining qty", "Seed", datetime.now()],
            ["number of", "Remaining qty", "Seed", datetime.now()],
            ["inventory level", "Remaining qty", "Seed", datetime.now()],
            ["units available", "Remaining qty", "Seed", datetime.now()],
            ["current stock", "Remaining qty", "Seed", datetime.now()],
            ["how many units", "Remaining qty", "Seed", datetime.now()],
            ["remaining balance", "Remaining qty", "Seed", datetime.now()],
            ["balance remaining", "Remaining qty", "Seed", datetime.now()],
            ["pieces left", "Remaining qty", "Seed", datetime.now()],
            ["stock count", "Remaining qty", "Seed", datetime.now()],
            ["availability", "Remaining qty", "Seed", datetime.now()],
            ["how many left", "Remaining qty", "Seed", datetime.now()],
            ["remaining quantity", "Remaining qty", "Seed", datetime.now()],
            ["stock left", "Remaining qty", "Seed", datetime.now()],
            ["quantity remaining", "Remaining qty", "Seed", datetime.now()],
            ["still in stock", "Remaining qty", "Seed", datetime.now()],
            ["units left", "Remaining qty", "Seed", datetime.now()],
            ["pieces remaining", "Remaining qty", "Seed", datetime.now()],
            ["how many still available", "Remaining qty", "Seed", datetime.now()],
            ["how many in inventory", "Remaining qty", "Seed", datetime.now()],
            
            # Price synonyms
            ["what's the price", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["how much is", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["what is the price of", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["net price", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["dealer price", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["cost of", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["price point", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["price per unit", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["unit price", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["cost per item", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["what does it cost", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["pricing for", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["quote for", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["contract price", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["quoted price", "Dealer net price \n[USD]", "Seed", datetime.now()],
            ["going for", "Dealer net price \n[USD]", "Seed", datetime.now()],
            
            # Family synonyms
            ["what family is", "Product family", "Seed", datetime.now()],
            ["product type", "Product family", "Seed", datetime.now()],
            ["what kind of", "Product family", "Seed", datetime.now()],
            ["what type of", "Product family", "Seed", datetime.now()],
            ["product line", "Product family", "Seed", datetime.now()],
            ["product category", "Product family", "Seed", datetime.now()],
            ["device type", "Product family", "Seed", datetime.now()],
            ["model family", "Product family", "Seed", datetime.now()],
            ["equipment type", "Product family", "Seed", datetime.now()],
            ["item category", "Product family", "Seed", datetime.now()],
            ["device category", "Product family", "Seed", datetime.now()],
            ["product series", "Product family", "Seed", datetime.now()],
            ["product classification", "Product family", "Seed", datetime.now()],
            ["what series is", "Product family", "Seed", datetime.now()],
            ["family of", "Product family", "Seed", datetime.now()],
            
            # Customer synonyms
            ["who is the customer", "Customer", "Seed", datetime.now()],
            ["end user", "Customer", "Seed", datetime.now()],
            ["who owns", "Customer", "Seed", datetime.now()],
            ["belongs to", "Customer", "Seed", datetime.now()],
            ["account holder", "Customer", "Seed", datetime.now()],
            ["customer name", "Customer", "Seed", datetime.now()],
            ["assigned to", "Customer", "Seed", datetime.now()],
            ["deal owner", "Customer", "Seed", datetime.now()],
            ["purchaser", "Customer", "Seed", datetime.now()],
            ["buyer name", "Customer", "Seed", datetime.now()],
            ["client account", "Customer", "Seed", datetime.now()],
            ["belongs to whom", "Customer", "Seed", datetime.now()],
            ["customer account", "Customer", "Seed", datetime.now()],
            ["who bought this", "Customer", "Seed", datetime.now()],
            ["purchased by", "Customer", "Seed", datetime.now()],
            ["client name", "Customer", "Seed", datetime.now()],
            
            # End date synonyms
            ["when does it expire", "End date", "Seed", datetime.now()],
            ["expiration date", "End date", "Seed", datetime.now()],
            ["valid until", "End date", "Seed", datetime.now()],
            ["good through", "End date", "Seed", datetime.now()],
            ["valid through", "End date", "Seed", datetime.now()],
            ["good through date", "End date", "Seed", datetime.now()],
            ["date of expiration", "End date", "Seed", datetime.now()],
            ["expiry date", "End date", "Seed", datetime.now()],
            ["expires on", "End date", "Seed", datetime.now()],
            ["termination date", "End date", "Seed", datetime.now()],
            ["end of validity", "End date", "Seed", datetime.now()],
            ["active until", "End date", "Seed", datetime.now()],
            ["until when", "End date", "Seed", datetime.now()],
            ["when is it valid until", "End date", "Seed", datetime.now()],
            ["when does the deal end", "End date", "Seed", datetime.now()],
            ["closing date", "End date", "Seed", datetime.now()]
        ]
        
        # Add essential synonyms that don't already exist
        added = 0
        seed_synonyms = {}
        try:
            with open(csv_path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                for syn in essential_synonyms:
                    if syn[0].lower() not in existing_synonyms:
                        writer.writerow(syn)
                        
                        # Also add to seed dictionary
                        seed_synonyms[syn[0].lower()] = syn[1]
                        
                        # Add to current brain
                        if syn[0].lower() not in self.synonym_brain:
                            self.synonym_brain[syn[0].lower()] = syn[1]
                            
                        added += 1
                        existing_synonyms.add(syn[0].lower())  # Update set to avoid duplicates
            
            # Add to sources
            if "Seed" not in self.brain_sources:
                self.brain_sources["Seed"] = {}
                
            self.brain_sources["Seed"].update(seed_synonyms)
            
            if added > 0:
                logger.info(f"✅ Seeded {added} essential synonyms to {csv_path}")
                if "Seed" not in [s.split(':')[0] for s in self.loaded_sources]:
                    self.loaded_sources.append(f"Seed:{added}")
                return True
            else:
                logger.info(f"✓ No new essential synonyms needed (all already exist)")
                return True
        except Exception as e:
            logger.error(f"❌ Failed to seed synonyms: {e}")
            self.metrics.record_error("Seed", e, {"path": csv_path})
            return False
            
    def load_product_family_map(self, excel_path: str = VARIABLES_FILE) -> int:
        """
        Load product family mapping with flexible sheet detection
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            int: Number of mappings loaded
        """
        if not HAS_OPENPYXL:
            logger.error("❌ openpyxl not available - cannot load product family map")
            return 0
            
        if not os.path.exists(excel_path):
            logger.error(f"❌ Excel file not found for product family map: {excel_path}")
            return 0
            
        try:
            start_time = time.time()
            
            # Try to load workbook
            wb = load_workbook(excel_path, read_only=True)
            
            # Define product family sheet variations
            family_variations = [
                "product family", "productfamily", "product_family", 
                "families", "family", "product families",
                "family map", "family mapping", "product map",
                "product mapping", "family types", "product types",
                "products", "product list", "items", "item families",
                "classification", "categories", "products by family"
            ]
            
            # Find the product family sheet
            sheet_name = self.find_matching_sheet(wb, family_variations)
            
            # Process Product family sheet if found
            family_count = 0
            if sheet_name:
                ws = wb[sheet_name]
                logger.info(f"✅ Found Product family sheet: '{sheet_name}'")
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 2 and row[0] and row[1]:
                        self.product_family_map[str(row[0]).strip().lower()] = str(row[1]).strip()
                        family_count += 1
                
                elapsed = time.time() - start_time
                
                logger.info(f"✅ Loaded {family_count} product family mappings in {elapsed:.2f}s")
                
                # Record metrics
                self.metrics.record_load_time("ProductFamily", elapsed, family_count)
                
                return family_count
            else:
                logger.warning(f"⚠️ Product family sheet not found in {excel_path}")
                return 0
                
        except Exception as e:
            logger.error(f"❌ Failed to load Product family map: {type(e).__name__}: {e}")
            self.metrics.record_error("ProductFamily", e, {"path": excel_path})
            return 0
            
    def log_brain_statistics(self) -> None:
        """Log statistics about the loaded brain"""
        try:
            # Count synonyms by field
            field_counts = {}
            for _, field in self.synonym_brain.items():
                field_counts[field] = field_counts.get(field, 0) + 1
                
            # Count synonyms by source and field
            source_field_counts = {}
            for source, synonyms in self.brain_sources.items():
                source_field_counts[source] = {}
                for _, field in synonyms.items():
                    source_field_counts[source][field] = source_field_counts[source].get(field, 0) + 1
                
            # Log summary
            logger.info(f"=== Brain Statistics ===")
            logger.info(f"Total synonyms: {len(self.synonym_brain)}")
            logger.info(f"Total fields: {len(field_counts)}")
            logger.info(f"Loaded from: {', '.join(self.loaded_sources)}")
            
            # Log count by field
            logger.info(f"Synonym counts by field:")
            for field, count in sorted(field_counts.items(), key=lambda x: x[1], reverse=True):
                logger.info(f"  - {field}: {count} synonyms")
                
            # Log count by source and field
            logger.info(f"Synonym counts by source and field:")
            for source, counts in sorted(source_field_counts.items()):
                logger.info(f"  {source}:")
                for field, count in sorted(counts.items(), key=lambda x: x[1], reverse=True):
                    logger.info(f"    - {field}: {count} synonyms")
                
            # Log a few examples from each source
            logger.info(f"Sample entries by source:")
            for source, synonyms in self.brain_sources.items():
                sample_count = min(3, len(synonyms))
                logger.info(f"  {source} samples:")
                for synonym, field in list(synonyms.items())[:sample_count]:
                    logger.info(f"    - '{synonym}' -> '{field}'")
                
            # Debug log - write full brain to debug file
            with open(DEBUG_LOG_FILE, 'w') as f:
                f.write(f"Brain loaded at: {datetime.now()}\n")
                f.write(f"Version: {VERSION}\n")
                f.write(f"Total synonyms: {len(self.synonym_brain)}\n")
                f.write(f"Sources: {', '.join(self.loaded_sources)}\n\n")
                
                f.write("=== SYNONYM MAP ===\n\n")
                
                # Output by field
                for field in sorted(field_counts.keys()):
                    f.write(f"--- {field} ---\n")
                    field_synonyms = sorted([k for k, v in self.synonym_brain.items() if v == field])
                    for syn in field_synonyms:
                        f.write(f"  {syn}\n")
                    f.write("\n")
        except Exception as e:
            logger.error(f"❌ Error logging brain statistics: {e}")
            
    def load_all(self) -> bool:
        """
        Load the synonym brain using a hybrid approach:
        1. Start with Golden Brain (highest priority)
        2. Add hardcoded brain for completeness
        3. Add Excel-based brain if available
        4. Add learned synonyms from CSV
        5. Add custom synonyms
        
        Returns:
            bool: Success or failure
        """
        overall_start = time.time()
        logger.info(f"=== Brain Loader v{VERSION} ===")
        logger.info(f"Starting hybrid brain loading process")
        
        # Reset state for reload
        self.synonym_brain = {}
        self.brain_sources = {}
        self.loaded_sources = []
        self.error_count = 0
        
        # Step 1: Start with Golden Brain (highest priority)
        logger.info(f"Step 1/7: Loading Golden Brain")
        self.load_golden_brain()
        
        # Step 2: Add hardcoded brain for completeness
        logger.info(f"Step 2/7: Adding Hardcoded Brain")
        self.load_hardcoded_brain()
        
        # Step 3: Try to load from Excel (primary external source)
        logger.info(f"Step 3/7: Loading from Excel: {VARIABLES_FILE}")
        self.load_from_excel(VARIABLES_FILE)
        
        # Step 4: Always try to load from CSV (learned synonyms)
        logger.info(f"Step 4/7: Loading from Learned CSV: {LEARNED_SYNONYMS_FILE}")
        self.load_from_csv(LEARNED_SYNONYMS_FILE, "LearnedCSV")
        
        # Step 5: Try to load from custom synonyms if available
        logger.info(f"Step 5/7: Loading from Custom CSV: {CUSTOM_SYNONYMS_FILE}")
        self.load_from_csv(CUSTOM_SYNONYMS_FILE, "CustomCSV")
        
        # Step 6: Try to seed essential synonyms if needed
        logger.info(f"Step 6/7: Checking for essential synonyms")
        self.seed_learned_synonyms(LEARNED_SYNONYMS_FILE)
        
        # Step 7: Try to load product family map
        logger.info(f"Step 7/7: Loading product family map")
        self.load_product_family_map(VARIABLES_FILE)
        
        # Save a backup of the successfully loaded brain
        if len(self.synonym_brain) > 0:
            self.save_brain_backup()
            
        # Log final statistics
        total_loaded = len(self.synonym_brain)
        logger.info(f"Step 8/8: Logging brain statistics")
        self.log_brain_statistics()
        
        # Save metrics
        self.metrics.save_metrics()
        
        # Summarize coverage by field
        field_counts = {}
        for _, field in self.synonym_brain.items():
            field_counts[field] = field_counts.get(field, 0) + 1
            
        logger.info("Field coverage summary:")
        for field in self.canonical_fields:
            count = field_counts.get(field, 0)
            logger.info(f"  - {field}: {count} synonyms")
        
        # Summary
        self.last_load_time = datetime.now()
        overall_elapsed = time.time() - overall_start
        
        if total_loaded > 0:
            logger.info(f"✅ Brain loading complete in {overall_elapsed:.2f}s. Total synonyms: {total_loaded} from {len(self.loaded_sources)} sources")
            return True
        else:
            logger.error(f"❌ CRITICAL: Failed to load any synonyms after {overall_elapsed:.2f}s!")
            return False
    
    def add_synonym(self, phrase: str, field: str, source: str = "Manual", confidence: int = 100) -> bool:
        """
        Add a new synonym and save it to the learned synonyms file
        
        Args:
            phrase: The synonym phrase
            field: The canonical field it maps to
            source: Source of the synonym
            confidence: Confidence score
            
        Returns:
            bool: Success or failure
        """
        if not phrase or not field:
            logger.warning(f"⚠️ Cannot add empty synonym: '{phrase}' -> '{field}'")
            return False
            
        # Validate field is canonical
        if field not in self.canonical_fields:
            logger.warning(f"⚠️ Cannot add synonym with non-canonical field: '{field}'")
            return False
            
        phrase_lower = phrase.lower().strip()
        
        # Check if it already exists with same mapping
        if phrase_lower in self.synonym_brain and self.synonym_brain[phrase_lower] == field:
            logger.info(f"✓ Synonym already exists: '{phrase}' -> '{field}'")
            return True
            
        # Add to brain
        self.synonym_brain[phrase_lower] = field
        
        # Add to appropriate source in brain_sources
        if source not in self.brain_sources:
            self.brain_sources[source] = {}
            
        self.brain_sources[source][phrase_lower] = field
        
        # Add to CSV
        try:
            with open(LEARNED_SYNONYMS_FILE, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                timestamp = datetime.now().isoformat()
                writer.writerow([phrase, field, source, timestamp, confidence])
                
            logger.info(f"✅ Added new synonym: '{phrase}' -> '{field}' (Source: {source}, Confidence: {confidence})")
            return True
        except Exception as e:
            logger.error(f"❌ Failed to save new synonym: {e}")
            # Still keep it in memory even if saving to CSV failed
            return True
            
    def get_synonym(self, phrase: str, default: Optional[str] = None) -> Optional[str]:
        """
        Look up a phrase in the synonym brain
        
        Args:
            phrase: The phrase to look up
            default: Default value if not found
            
        Returns:
            str: The field associated with the phrase, or default if not found
        """
        if not phrase:
            return default
            
        phrase_lower = phrase.lower().strip()
        return self.synonym_brain.get(phrase_lower, default)
    
    def has_field(self, text: str, field: str) -> bool:
        """
        Check if text contains any synonyms for the specified field
        
        Args:
            text: Text to check
            field: Field to check for
            
        Returns:
            bool: True if text contains synonyms for field
        """
        if not text or not field:
            return False
            
        text_lower = text.lower()
        
        # Direct synonym check
        for synonym, mapped_field in self.synonym_brain.items():
            if mapped_field == field and re.search(r'\b' + re.escape(synonym) + r'\b', text_lower):
                return True
                
        return False
    
    def get_field_coverage(self) -> Dict[str, int]:
        """
        Get coverage statistics for each field
        
        Returns:
            dict: Field coverage statistics
        """
        field_counts = {}
        
        # Initialize with zeros for all canonical fields
        for field in self.canonical_fields:
            field_counts[field] = 0
            
        # Count occurrences
        for _, field in self.synonym_brain.items():
            if field in field_counts:
                field_counts[field] += 1
                
        return field_counts
    
    def get_source_coverage(self) -> Dict[str, Dict[str, int]]:
        """
        Get coverage statistics by source and field
        
        Returns:
            dict: Source and field coverage statistics
        """
        coverage = {}
        
        for source, synonyms in self.brain_sources.items():
            coverage[source] = {}
            
            # Initialize with zeros for all canonical fields
            for field in self.canonical_fields:
                coverage[source][field] = 0
                
            # Count occurrences
            for _, field in synonyms.items():
                if field in coverage[source]:
                    coverage[source][field] += 1
                    
        return coverage
    
    def validate_brain(self) -> Tuple[bool, Dict[str, Any]]:
        """
        Validate the brain for completeness and consistency
        
        Returns:
            tuple: (is_valid, validation_results)
        """
        validation_results = {
            "timestamp": datetime.now().isoformat(),
            "total_synonyms": len(self.synonym_brain),
            "field_coverage": self.get_field_coverage(),
            "source_coverage": self.get_source_coverage(),
            "issues": [],
            "warnings": [],
            "status": "valid"
        }
        
        # Check minimum thresholds for each field
        min_thresholds = {
            "Remaining qty": 20,
            "Dealer net price \n[USD]": 20,
            "Product family": 20,
            "Customer": 20,
            "End date": 20
        }
        
        field_coverage = validation_results["field_coverage"]
        
        for field, threshold in min_thresholds.items():
            count = field_coverage.get(field, 0)
            if count < threshold:
                validation_results["issues"].append({
                    "type": "low_coverage",
                    "field": field,
                    "count": count,
                    "threshold": threshold
                })
                validation_results["status"] = "warning"
                
        # Check for non-canonical fields
        non_canonical = [field for field in field_coverage if field not in self.canonical_fields]
        if non_canonical:
            validation_results["issues"].append({
                "type": "non_canonical_fields",
                "fields": non_canonical
            })
            validation_results["status"] = "warning"
            
        # Check source diversity
        if len(self.brain_sources) < 2:
            validation_results["warnings"].append({
                "type": "low_source_diversity",
                "count": len(self.brain_sources)
            })
            
        # Critical check - any fields with zero coverage?
        zero_coverage = [field for field in self.canonical_fields if field_coverage.get(field, 0) == 0]
        if zero_coverage:
            validation_results["issues"].append({
                "type": "zero_coverage",
                "fields": zero_coverage
            })
            validation_results["status"] = "invalid"
            
        # Check for any empty sources
        empty_sources = [source for source, synonyms in self.brain_sources.items() if not synonyms]
        if empty_sources:
            validation_results["warnings"].append({
                "type": "empty_sources",
                "sources": empty_sources
            })
            
        # Set overall validity
        is_valid = validation_results["status"] != "invalid"
        
        return is_valid, validation_results
    
    def diagnostic_report(self) -> Dict[str, Any]:
        """Generate a diagnostic report of the brain state"""
        is_valid, validation = self.validate_brain()
        
        return {
            "version": VERSION,
            "last_updated": LAST_UPDATED,
            "last_load_time": self.last_load_time.isoformat() if self.last_load_time else None,
            "brain_size": len(self.synonym_brain),
            "sources": self.loaded_sources,
            "error_count": self.error_count,
            "field_coverage": self.get_field_coverage(),
            "source_coverage": self.get_source_coverage(),
            "is_valid": is_valid,
            "validation": validation
        }

# Global instance
brain_loader = BrainLoader()

# Wrapper functions for backward compatibility

def load_synonym_brain() -> Dict[str, str]:
    """
    Wrapper function for backward compatibility
    
    Returns:
        dict: The loaded synonym brain
    """
    global brain_loader
    brain_loader.load_all()
    return brain_loader.synonym_brain

def load_product_family_map() -> Dict[str, str]:
    """
    Wrapper function for backward compatibility
    
    Returns:
        dict: The loaded product family map
    """
    global brain_loader
    if not brain_loader.product_family_map:
        brain_loader.load_product_family_map()
    return brain_loader.product_family_map

def add_synonym(phrase: str, field: str, source: str = "Manual", confidence: int = 100) -> bool:
    """
    Wrapper function for backward compatibility
    
    Args:
        phrase: The synonym phrase
        field: The field it maps to
        source: Source of this synonym
        confidence: Confidence score
        
    Returns:
        bool: Success or failure
    """
    global brain_loader
    return brain_loader.add_synonym(phrase, field, source, confidence)

def debug_excel_structure(excel_path: str = VARIABLES_FILE) -> None:
    """Debug helper to print Excel file structure"""
    if not HAS_OPENPYXL:
        logger.error("❌ openpyxl not available - cannot debug Excel structure")
        return
        
    logger.info(f"Analyzing Excel structure: {excel_path}")
    
    try:
        wb = load_workbook(excel_path, read_only=True)
        logger.info(f"Excel file: {excel_path}")
        logger.info(f"Sheets: {wb.sheetnames}")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get dimensions
            min_row, max_row = ws.min_row, ws.max_row
            if max_row < 2:  # Empty sheet
                logger.info(f"Sheet: '{sheet_name}' - Empty (no data rows)")
                continue
                
            # Sample the first row to determine format
            header_row = next(ws.rows)
            header_values = [cell.value for cell in header_row]
            
            # Count non-empty header cells
            non_empty_headers = sum(1 for h in header_values if h)
            
            logger.info(f"Sheet: '{sheet_name}', Rows: {max_row}, Headers: {non_empty_headers}")
            
            # Print first few rows as sample
            sample_count = min(3, max_row)
            logger.info(f"Sample data from '{sheet_name}':")
            
            for i, row in enumerate(list(ws.rows)[:sample_count]):
                values = [str(cell.value) if cell.value is not None else "" for cell in row]
                sample = ", ".join(values[:5])  # Show first 5 columns
                if len(values) > 5:
                    sample += ", ..."
                logger.info(f"  Row {i+1}: {sample}")
    except Exception as e:
        logger.error(f"Error debugging Excel: {type(e).__name__}: {e}")
        logger.error(traceback.format_exc())

# Run when executed directly
if __name__ == "__main__":
    logger.info("Starting brain_loader_v5.0 diagnostic")
    
    print(f"Brain Loader v{VERSION}")
    print(f"Last updated: {LAST_UPDATED}")
    print("\nRunning full diagnostic...")
    
    # Start with Excel structure analysis
    if HAS_OPENPYXL and os.path.exists(VARIABLES_FILE):
        print("\nAnalyzing Excel structure...")
        debug_excel_structure()
    
    # Load the brain with comprehensive approach
    print("\nLoading brain with hybrid approach...")
    start_time = time.time()
    brain = load_synonym_brain()
    elapsed = time.time() - start_time
    
    # Print summary
    print(f"\nLoaded {len(brain)} synonyms in {elapsed:.2f} seconds")
    print(f"Sources: {', '.join(brain_loader.loaded_sources)}")
    
    # Print field coverage
    field_coverage = brain_loader.get_field_coverage()
    print("\nField coverage:")
    for field, count in sorted(field_coverage.items(), key=lambda x: x[1], reverse=True):
        print(f"  {field}: {count} synonyms")
    
    # Validate brain
    is_valid, validation = brain_loader.validate_brain()
    print(f"\nBrain validation: {'✅ Valid' if is_valid else '❌ Invalid'}")
    
    if validation["issues"]:
        print("\nIssues detected:")
        for issue in validation["issues"]:
            print(f"  - {issue['type']}: {issue}")
    
    if validation["warnings"]:
        print("\nWarnings:")
        for warning in validation["warnings"]:
            print(f"  - {warning['type']}: {warning}")
    
    # Test key phrases
    test_phrases = [
        "how many left", "what is the price", "when does it expire", 
        "who is the customer", "what family is this", "quantity remaining",
        "units available", "cost of", "expiration date"
    ]
    
    print("\nTesting key phrases:")
    for phrase in test_phrases:
        field = brain_loader.get_synonym(phrase)
        result = f"'{field}'" if field else "Not found"
        print(f"  '{phrase}' -> {result}")
    
    # Generate and display metrics
    print("\nMetrics summary:")
    metrics = brain_loader.metrics.get_summary()
    for key, value in metrics.items():
        if not isinstance(value, dict) and not isinstance(value, list):
            print(f"  {key}: {value}")
    
    print("\nDiagnostics complete.")